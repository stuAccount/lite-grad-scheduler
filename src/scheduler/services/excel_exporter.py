"""Excel export service for course data."""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from scheduler.domain.models import Course


class ScheduleExcelExporter:
    """Generate Excel schedules."""
    
    def generate_course_list(self, courses: list[Course]) -> BytesIO:
        """Generate Excel file with course list.
        
        Args:
            courses: List of all courses
            
        Returns:
            BytesIO buffer containing Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Course Schedule"
        
        # Headers
        headers = [
            'Course ID', 'Course Name', 'Professor ID', 'Classroom ID',
            'Day', 'Period', 'Credits', 'Hours', 'Type', 'Department'
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Weekday mapping
        weekday_names = {
            1: 'Monday', 2: 'Tuesday', 3: 'Wednesday',
            4: 'Thursday', 5: 'Friday'
        }
        
        # Data rows (sorted by weekday and period)
        for course in sorted(courses, key=lambda c: (c.weekday or 0, c.period or 0)):
            ws.append([
                course.id,
                course.name,
                course.professor_id,
                course.classroom_id,
                weekday_names.get(course.weekday, '') if course.weekday else '',
                course.period if course.period else '',
                course.credits if course.credits else '',
                course.hours if course.hours else '',
                course.course_type if course.course_type else '',
                course.department if course.department else '',
            ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
